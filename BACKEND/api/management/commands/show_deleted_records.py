"""
Management command to display and export deleted records from the backup table.
Backend-only access (not exposed in UI).

Usage:
    python manage.py show_deleted_records                          # Show all deleted records
    python manage.py show_deleted_records --type=user              # Show only deleted users
    python manage.py show_deleted_records --type=product           # Show only deleted products
    python manage.py show_deleted_records --type=sale              # Show only deleted sales
    python manage.py show_deleted_records --limit=50               # Show last 50 records
    python manage.py show_deleted_records --json                   # Output as JSON
    python manage.py show_deleted_records --export=sql             # Export as SQL INSERT file
    python manage.py show_deleted_records --export=excel           # Export as Excel (.xlsx) file
    python manage.py show_deleted_records --export=csv             # Export as CSV file
    python manage.py show_deleted_records --type=product --export=sql  # Export only products as SQL
"""

from django.core.management.base import BaseCommand
from django.utils import timezone
from api.models import DeletedRecordsBackup
import json
import os


class Command(BaseCommand):
    help = 'Display and export deleted records from backup table (backend access only)'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--type',
            type=str,
            default=None,
            help='Filter by record type (user, product, sale)',
        )
        parser.add_argument(
            '--limit',
            type=int,
            default=20,
            help='Number of records to display (default: 20)',
        )
        parser.add_argument(
            '--json',
            action='store_true',
            help='Output as JSON to terminal',
        )
        parser.add_argument(
            '--export',
            type=str,
            default=None,
            choices=['sql', 'excel', 'csv'],
            help='Export format: sql, excel, or csv',
        )
        parser.add_argument(
            '--output',
            type=str,
            default=None,
            help='Output file path (auto-generated if not specified)',
        )
    
    def _build_queryset(self, options):
        """Build and return the filtered queryset."""
        record_type = options.get('type')
        limit = options.get('limit')
        
        queryset = DeletedRecordsBackup.objects.all()
        
        if record_type:
            queryset = queryset.filter(record_type=record_type)
        
        queryset = queryset.order_by('-deleted_at')[:limit]
        return queryset
    
    def _get_output_path(self, options, extension):
        """Generate output file path."""
        output = options.get('output')
        if output:
            return output
        
        record_type = options.get('type') or 'all'
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'deleted_records_{record_type}_{timestamp}.{extension}'
        
        # Save to user's Downloads folder
        downloads_dir = os.path.join(os.path.expanduser('~'), 'Downloads')
        os.makedirs(downloads_dir, exist_ok=True)
        return os.path.join(downloads_dir, filename)
    
    def _export_sql(self, queryset, options):
        """Export as SQL INSERT statements."""
        output_path = self._get_output_path(options, 'sql')
        
        lines = []
        lines.append('-- Deleted Records Backup Export')
        lines.append(f'-- Generated: {timezone.now().isoformat()}')
        lines.append(f'-- Records: {queryset.count()}')
        lines.append('')
        lines.append('-- Table: api_deletedrecordsbackup')
        lines.append('-- Columns: id, record_type, original_id, data, deleted_at, deleted_by_id')
        lines.append('')
        
        for record in queryset:
            data_json = json.dumps(record.data, default=str).replace("'", "''")
            deleted_by_id = record.deleted_by_id if record.deleted_by_id else 'NULL'
            deleted_at = record.deleted_at.strftime('%Y-%m-%d %H:%M:%S')
            
            lines.append(
                f"INSERT INTO api_deletedrecordsbackup (id, record_type, original_id, data, deleted_at, deleted_by_id) VALUES "
                f"({record.id}, '{record.record_type}', {record.original_id}, "
                f"'{data_json}', '{deleted_at}', {deleted_by_id});"
            )
        
        lines.append('')
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        self.stdout.write(self.style.SUCCESS(f'\n[OK] SQL exported: {output_path}'))
        self.stdout.write(f'   Records: {queryset.count()}')
        self.stdout.write(f'   Format: SQL INSERT statements\n')
    
    def _export_excel(self, queryset, options):
        """Export as Excel (.xlsx) file."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.stdout.write(self.style.ERROR(
                '\n[ERROR] openpyxl is required for Excel export.\n'
                '   Install it: pip install openpyxl\n'
            ))
            return
        
        output_path = self._get_output_path(options, 'xlsx')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Deleted Records'
        
        # --- Styles ---
        header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='4A2F21', end_color='4A2F21', fill_type='solid')  # espresso
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # --- Headers ---
        headers = ['ID', 'Type', 'Original ID', 'Deleted At', 'Deleted By', 'Name/Identifier', 'Full Data (JSON)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # --- Data Rows ---
        for row_idx, record in enumerate(queryset, 2):
            # Extract a name/identifier from the data
            name = ''
            if record.data:
                name = record.data.get('name', '') or record.data.get('username', '') or record.data.get('order_id', '') or str(record.original_id)
            
            values = [
                record.id,
                record.record_type,
                record.original_id,
                record.deleted_at.strftime('%Y-%m-%d %H:%M:%S'),
                record.deleted_by.username if record.deleted_by else 'Unknown',
                name,
                json.dumps(record.data, indent=2, default=str) if record.data else '',
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        # --- Column Widths ---
        ws.column_dimensions['A'].width = 6    # ID
        ws.column_dimensions['B'].width = 10   # Type
        ws.column_dimensions['C'].width = 12   # Original ID
        ws.column_dimensions['D'].width = 22   # Deleted At
        ws.column_dimensions['E'].width = 15   # Deleted By
        ws.column_dimensions['F'].width = 25   # Name
        ws.column_dimensions['G'].width = 60   # Full Data
        
        # --- Freeze header row ---
        ws.freeze_panes = 'A2'
        
        # --- Auto-filter ---
        ws.auto_filter.ref = f'A1:G{queryset.count() + 1}'
        
        wb.save(output_path)
        
        self.stdout.write(self.style.SUCCESS(f'\n[OK] Excel exported: {output_path}'))
        self.stdout.write(f'   Records: {queryset.count()}')
        self.stdout.write(f'   Format: Excel (.xlsx) with filters and styling\n')
    
    def _export_csv(self, queryset, options):
        """Export as CSV file."""
        import csv
        
        output_path = self._get_output_path(options, 'csv')
        
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow(['ID', 'Type', 'Original ID', 'Deleted At', 'Deleted By', 'Name/Identifier', 'Full Data (JSON)'])
            
            # Rows
            for record in queryset:
                name = ''
                if record.data:
                    name = record.data.get('name', '') or record.data.get('username', '') or record.data.get('order_id', '') or str(record.original_id)
                
                writer.writerow([
                    record.id,
                    record.record_type,
                    record.original_id,
                    record.deleted_at.strftime('%Y-%m-%d %H:%M:%S'),
                    record.deleted_by.username if record.deleted_by else 'Unknown',
                    name,
                    json.dumps(record.data, default=str) if record.data else '',
                ])
        
        self.stdout.write(self.style.SUCCESS(f'\n[OK] CSV exported: {output_path}'))
        self.stdout.write(f'   Records: {queryset.count()}')
        self.stdout.write(f'   Format: CSV (UTF-8 with BOM, Excel-compatible)\n')
    
    def handle(self, *args, **options):
        queryset = self._build_queryset(options)
        
        if not queryset.exists():
            self.stdout.write(self.style.WARNING('No deleted records found.'))
            return
        
        # --- Export mode ---
        export_format = options.get('export')
        if export_format == 'sql':
            self._export_sql(queryset, options)
            return
        elif export_format == 'excel':
            self._export_excel(queryset, options)
            return
        elif export_format == 'csv':
            self._export_csv(queryset, options)
            return
        
        # --- Terminal display mode ---
        as_json = options.get('json')
        
        if as_json:
            records = []
            for record in queryset:
                records.append({
                    'id': record.id,
                    'record_type': record.record_type,
                    'original_id': record.original_id,
                    'data': record.data,
                    'deleted_at': record.deleted_at.isoformat(),
                    'deleted_by': record.deleted_by.username if record.deleted_by else None,
                })
            self.stdout.write(json.dumps(records, indent=2, default=str))
        else:
            self.stdout.write(self.style.SUCCESS(f'\n=== Deleted Records Backup ({queryset.count()} records) ===\n'))
            
            for record in queryset:
                self.stdout.write(self.style.SUCCESS(f'ID: {record.id}'))
                self.stdout.write(f'  Type: {record.record_type}')
                self.stdout.write(f'  Original ID: {record.original_id}')
                self.stdout.write(f'  Deleted At: {record.deleted_at}')
                self.stdout.write(f'  Deleted By: {record.deleted_by.username if record.deleted_by else "Unknown"}')
                self.stdout.write(f'  Data:')
                
                try:
                    for key, value in record.data.items():
                        self.stdout.write(f'    {key}: {value}')
                except:
                    self.stdout.write(f'    {record.data}')
                
                self.stdout.write('')
